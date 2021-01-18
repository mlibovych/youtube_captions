from bs4 import BeautifulSoup
import lxml
import requests
from pytube import YouTube
from selenium import webdriver
import time
from dbclass import Database
import openpyxl
from openpyxl import load_workbook
import os
from threading import Thread
import threading

search_url = 'https://www.youtube.com/results?search_query='
youtube_url = 'https://www.youtube.com'
database = Database()
files_directory = 'caption_files'
captions_table = 'captions.xlsx'

DONE = 1
NOT_DONE = 0

if os.name == 'posix':
    delimeter = '/'
else:
    delimeter = '\\'


class Resp:
    def __init__(self):
        self.source = None


def create_youtube_object(resp, url):
    resp.source = YouTube(url)


class YouTubeClass:
    def __init__(self, film):
        film_words = film.split(' ')
        film_words.append('movie')
        film_words.append('review')
        film_search_url = search_url + "+".join(film_words)  # получаем юрл поиска

        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')

        self.browser = webdriver.Chrome(os.path.dirname(os.path.abspath(__file__)) + delimeter + 'chromedriver', options=chrome_options)
        self.browser.get(film_search_url)
        self.request = 1
        time.sleep(2)

        # считаем количество совпадений
        videos = self.browser.find_elements_by_tag_name('ytd-video-renderer')
        count = 0

        for video in videos:
            title = video.find_element_by_id('video-title')

            if title.text.lower().find(film.lower()) != -1:
                count += 1

        # если совпадений меньше 5 меняем запрос
        if count < 5:
            new_url = search_url + "\"" + film + "\"" # получаем юрл поиска в ""

            self.browser.get(new_url)
            time.sleep(2)
            self.request = 2

        self.html = self.browser.page_source

    def get_html(self):
        return self.html

    def get_request(self):
        return self.request

    def close(self):
        self.browser.close()
        self.browser.quit()


def get_youtube_html(film):
    source = YouTubeClass(film)

    html = source.get_html()  # получаем код страницы
    request = source.get_request() # получаем номер запроса

    source.close()

    return html, request


def get_film_links(film):
    html, request = get_youtube_html(film) #получаем код страницы
    soup = BeautifulSoup(html, 'lxml')
    videos = soup.find_all('ytd-video-renderer') #получаем список видео

    links = []
    for video in videos:
        link = video.find('a', id='video-title')
        # if link.get('title').lower().find(film.lower()) != -1: #если в названии есть полное совпадение с фильмом
        links.append(link.get('href'))

    return links, request


def delete_timecodes(captions):
    split_captions = captions.split('\n')
    result = []
    count = 0

    for i in range(len(split_captions)):
        count += 1
        if count == 3:
            result.append(split_captions[i])
        if count == 4:
            count = 0

    clear_captions = "\n".join(result)

    return clear_captions


def write_in_file(film, captions, count):
    film = film.replace('/', '-')
    filename = files_directory + delimeter + film + '_' + str(count) + '.txt'

    with open(filename, 'w+') as file:
        file.write(captions)

    return filename


def get_human_readable_view(amount):
    human_readable = str(amount)
    if len(human_readable) < 2:
        human_readable = '0' + human_readable

    return human_readable


def get_duration(length):
    seconds = length % 60
    mins = int((length - seconds) / 60)
    duration = get_human_readable_view(mins) + ':' + get_human_readable_view(seconds)

    return duration


class Captions:
    def __init__(self, film_id, code, video_duration, request, position, text):
        self.film_id = film_id
        self.code = code
        self.video_duration = video_duration
        self.request = request
        self.position = position
        self.text = text

    def get_fim_id(self):
        return self.film_id

    def get_code(self):
        return self.code

    def get_duration(self):
        return self.video_duration

    def get_request(self):
        return self.request

    def get_position(self):
        return self.position

    def get_text(self):
        return self.text

def get_video_captions(film_id, url, request, count):
    print(url)

    # try to get responce
    responce = Resp()
    thread = Thread(target=create_youtube_object, args=(responce, url))
    thread.start()

    # time.sleep(5)
    for i in range(0, 10):
        time.sleep(0.5)
        if responce.source != None:
            break

    if responce.source == None:
        print('err')
        return -1, None

    print('done')
    # --

    # получаем коды субтитров
    captions = responce.source.caption_tracks

    caption_codes = []
    for caption in captions:
        caption_codes.append(caption.code)

    en_captions = [cap for cap in caption_codes if cap.find('en') != -1]
    en_captions.sort(reverse=True)

    if len(en_captions):
        en_caption_code = en_captions[0]
        en_caption = responce.source.captions[en_caption_code]
        en_caption_convert_to_srt = (en_caption.generate_srt_captions())
        en_caption_text = delete_timecodes(en_caption_convert_to_srt)
        duration = responce.source.length

        caption = Captions(film_id, en_caption_code, duration, request, count, en_caption_text)

        return 1, caption

    return 0, None


def get_all_en_captions(film_id, film_name):
    links, request = get_film_links(film_name)
    captions = []
    count = 1

    for link in links:
        url = youtube_url + link
        try:
            res, caption = get_video_captions(film_id, url, request, count)
            tries = 0

            while res < 0:
                res, caption = get_video_captions(film_id, url, request, count)
                tries += 1
                if tries > 5:
                    res = 0
                    break

            if res > 0:
                captions.append(caption)

            count += 1
            if count > 10:
                break
        except Exception:
            print("Unknown err")

    return captions

def get_sub_words_count(caption):
    text = caption.text
    count = 0

    lines = text.split('\n')
    for line in lines:
        words = line.split(' ')
        count += len(words)

    return count


def get_best(caption):
    res = 0
    words = get_sub_words_count(caption)

    if caption.video_duration >= 300 and words >= 200:
        res += 30
    if caption.code[0] == 'e':
        res += 20
    res += 10 - caption.position

    return res


def get_captions(film_id, film_name):
    captions = get_all_en_captions(film_id, film_name)
    count = 0

    captions.sort(key=get_best, reverse=True)
    for caption in captions[:5]:
        # запись субтитров в файл
        file_name = write_in_file(film_name, caption.text, count)
        duration = get_duration(caption.video_duration)

        # критическая секция для многопоточности !!!

        # запись в базу
        caption_id = database.insert_caption(film_id, caption.code, duration,
                                             caption.request, caption.position, file_name)[0]

        # запись в файл
        write_in_xlsx(caption_id, film_id, caption.code, duration,
                      caption.request, caption.position, file_name)
        count += 1

    database.update_film(film_id, DONE)


def open_xlsx(file_name):
    # открываем файл для записи
    out = openpyxl.Workbook()

    # таблица фильмов
    out.create_sheet(title='films', index=0)
    films_sheet = out['films']
    film_titles = ['id', 'title']
    col = 1

    for title in film_titles:
        films_sheet.cell(row=1, column=col).value = title
        col += 1

    # таблица субтитров
    out.create_sheet(title='captions', index=1)
    captions_sheet = out['captions']
    caption_titles = ['id', 'film_id', 'code', 'duration', 'responce', 'count', 'file_name']
    col = 1

    for title in caption_titles:
        captions_sheet.cell(row=1, column=col).value = title
        col += 1

    # открываем файл для чтения
    wb = load_workbook(file_name)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        sheet = wb[sheet_name]
        rows = sheet.max_row

        for i in range(1, rows + 1):
            cell = sheet.cell(row=i, column=1)
            film_title = cell.value

            # запись в базу
            film_id = database.insert_film(film_title)[0]

            # запись в файл
            film_row = films_sheet.max_row + 1
            id_cell = films_sheet.cell(row=film_row, column=1)
            id_cell.value = film_id

            film_cell = films_sheet.cell(row=film_row, column=2)
            film_cell.value = film_title

    out.save(captions_table)


def write_in_xlsx(id, film_id, code, video_duration, request, position, file_name):
    if not os.path.exists(captions_table):
        wb = openpyxl.Workbook()
    else:
        wb = load_workbook(captions_table)

    captions_sheet = wb['captions']
    row = captions_sheet.max_row + 1

    captions_sheet.cell(row=row, column=1).value = id
    captions_sheet.cell(row=row, column=2).value = film_id
    captions_sheet.cell(row=row, column=3).value = code
    captions_sheet.cell(row=row, column=4).value = video_duration
    captions_sheet.cell(row=row, column=5).value = request
    captions_sheet.cell(row=row, column=6).value = position
    captions_sheet.cell(row=row, column=7).value = file_name

    wb.save(captions_table)


def main():
    if not os.path.exists(files_directory):
        os.mkdir(files_directory)

    films = database.get_not_proceeded_films()

    for film_id, film_title in films:
        print(film_id, film_title)
        get_captions(film_id, film_title)


main()


# def main():
#     open_xlsx('unique.xlsx')


# lock = threading.Lock()
#
# lock.acquire()
# lock.release()


